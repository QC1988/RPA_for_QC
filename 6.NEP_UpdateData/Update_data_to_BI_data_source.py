# Update_data_to_BI_data_source
#!/usr/bin/python
# -*- coding: <utf-8> -*-

import os
from sys import path
import glob
import pandas as pd
import openpyxl
import xlwings as xw
import time
import configparser

import import_paras_NEP_inspection_data


pwd = os.getcwd()
father_path=os.path.abspath(os.path.dirname(pwd)+os.path.sep+".")
path.insert(0, father_path)

# define class to void the chars changed to lower chars
class myconf(configparser.ConfigParser):
    def __init__(self, defaults=None):
        configparser.ConfigParser.__init__(self, defaults=None)
    def optionxform(self, optionstr):
        return optionstr

# input object:NEP inspection data, NG data|excel file;Frequency:every day
# output object:BI data source
# config:
# [paras_for_NEP]


# 0.load ParamS
# control the files copy to the server or not
Copy_To_Online = True

# C:\Users\035203557\Desktop\kaizen_space\RPA\6.NEP_UpdateData
Local_workpath = import_paras_NEP_inspection_data.Local_workpath


BI_data_source_online_path = import_paras_NEP_inspection_data.BI_data_source_online_path
# UPS_Acceptance_Defect_NEP_FY19-.xlsx
BI_data_source_file_name = import_paras_NEP_inspection_data.BI_data_source_file_name


NEP_inspection_data_online_path = import_paras_NEP_inspection_data.NEP_inspection_data_online_path
# OMRON UPS INSPECTION REPORT-
NEP_inspeciton_file_name = import_paras_NEP_inspection_data.NEP_inspeciton_file_name


NEP_NG_data_online_path = import_paras_NEP_inspection_data.NEP_NG_data_online_path
# NGReport_
NEP_NG_data_file_name = import_paras_NEP_inspection_data.NEP_NG_data_file_name


NEP_SN_data_online_path = import_paras_NEP_inspection_data.NEP_SN_data_online_path
# 20
NEP_SN_data_file_name = import_paras_NEP_inspection_data.NEP_SN_data_file_name


NEP_RoHS_online_path = import_paras_NEP_inspection_data.NEP_RoHS_online_path
NEP_RoHS_file_name = import_paras_NEP_inspection_data.NEP_RoHS_file_name


BI_data_source_online_path_name = import_paras_NEP_inspection_data.BI_data_source_online_path_name
BI_data_source_local_path_name = import_paras_NEP_inspection_data.BI_data_source_local_path_name

NEP_inspection_data_online_path_name = import_paras_NEP_inspection_data.NEP_inspection_data_online_path_name
NEP_inspection_data_local_path_name = import_paras_NEP_inspection_data.NEP_inspection_data_local_path_name

NEP_NG_data_online_path_name = import_paras_NEP_inspection_data.NEP_NG_data_online_path_name
NEP_NG_data_local_path_name = import_paras_NEP_inspection_data.NEP_NG_data_local_path_name

NEP_SN_data_online_path_name=import_paras_NEP_inspection_data.NEP_SN_data_online_path_name
NEP_SN_data_local_path_name=import_paras_NEP_inspection_data.NEP_SN_data_local_path_name
# 
NEP_RoHS_online_path_name = import_paras_NEP_inspection_data.NEP_RoHS_online_path_name
NEP_RoHS_local_path_name = import_paras_NEP_inspection_data.NEP_RoHS_local_path_name

# copy command
copy_command_NEP_BI_source_data_from_online_to_local = 'copy' + ' ' + '"' + BI_data_source_online_path_name  + '"'  +  ' ' + '"' + Local_workpath +'"'
copy_command_NEP_BI_source_data_from_local_to_online = 'copy' + ' ' + '"' + BI_data_source_local_path_name  + '"'  +  ' ' + '"' + BI_data_source_online_path +'"'

copy_command_NEP_inspection_data_from_online_to_local = 'copy' + ' ' +  '"' + NEP_inspection_data_online_path_name + '"'   + ' ' + '"' + Local_workpath +'"'
# copy_command_NEP_inspection_data_from_local_to_online = 'copy' + ' ' + '"' + NEP_inspection_data_local_path_name  + '"'  +  ' ' + '"' + NEP_inspection_data_online_path +'"'

copy_command_NEP_NG_data_from_online_to_local = 'copy' + ' ' + '"' + NEP_NG_data_online_path_name  + '"'  +  ' ' + '"' + Local_workpath +'"'
# copy_command_NEP_NG_data_from_local_to_online = 'copy' + ' ' + '"' + NEP_NG_data_local_path_name  + '"'  +  ' ' + '"' + NEP_NG_data_online_path +'"'

copy_command_NEP_SN_data_from_online_to_local = 'copy' + ' ' + '"' + NEP_SN_data_online_path_name  + '"'  +  ' ' + '"' + Local_workpath +'"'
# copy_command_NEP_SN_data_from_local_to_online = 'copy' + ' ' + '"' + NEP_SN_data_local_path_name  + '"'  +  ' ' + '"' + NEP_SN_data_online_path+'"'

copy_command_NEP_RoHS_from_online_to_local = 'copy' + ' ' + '"' + NEP_RoHS_online_path_name  + '"'  +  ' ' + '"' + Local_workpath +'"'
copy_command_NEP_RoHS_from_local_to_online = 'copy' + ' ' + '"' + NEP_RoHS_local_path_name  + '"'  +  ' ' + '"' + NEP_RoHS_online_path +'"'

# ver.1.1 NO_NEP_NG_REPORT == False



# define fuctions to write xlsx, value = DataFrame type
# fuction 1 *not used
def write_excel_xlsx_append_openpyxl(source_excel_path_name, sheet_name, last_row, dataFrame_insert_into_source):
    index = len(dataFrame_insert_into_source)
    print("%d rows will be written into the file."%index)
    # wb = openpyxl.Workbook(path_name)
    wb = openpyxl.load_workbook(source_excel_path_name)
    sheet = wb.active
    sheet.title = sheet_name
    for i in range(0, index):
        for j in range(0, len(dataFrame_insert_into_source.iloc[i,:])):
            value = dataFrame_insert_into_source.iloc[i,j]
            sheet.cell(row=i+last_row+2, column=j+1, value=value)
    wb.save(source_excel_path_name)
    wb.close()
    print("sheet %s has been written successfully."%sheet_name)

# fuction 2 *used
def write_excel_xlsx_append_xlwings(source_excel_path_name, sheet_name, last_row, dataFrame_insert_into_source, hearder_true_false):
    print("")
    print("")
    print("================================================================")
    print('    Fuction of write_excel_xlsx_append_xlwings app open.')

    print(dataFrame_insert_into_source)
    index = len(dataFrame_insert_into_source)
    print("    %d rows will be written into the file."%index)
    app = xw.App(visible=False, add_book=False)
    wb = app.books.open(source_excel_path_name)
    sheet_name = sheet_name
    sht = wb.sheets[sheet_name]
    for i in range(0, index):
        for j in range(0, len(dataFrame_insert_into_source.iloc[i,:])):
            value = dataFrame_insert_into_source.iloc[i,j]
            sht.range(i+last_row+1+hearder_true_false, j+1).value = value
    wb.save(source_excel_path_name)
    wb.close()
    app.quit()
    print('    Fuction of write_excel_xlsx_append_xlwings app closed.')
    print("    sheet %s has been written successfully."%sheet_name)
    print("================================================================")
    print("")
    print("")

# define a fuction to compare the value of two columns in source data and the data will be inserted.
def compare_two_columns_source_insert( dataFrame_source_column_list, dataFrame_insert_column_list):
    print("Checking all reports has not been import to the source data.")
    for i in dataFrame_source_column_list:
        for j in dataFrame_insert_column_list:
            if i == j:
                print("Error, %s has been import to source data or something wrong in order No."%j)
                exit()
    print("confimation OK, There is no same content in two files.")
    

def main():
    THERE_ARE_NEP_NG_REPORTS_IN_WORKPATH = True
    THERE_ARE_NEW_NEP_NG_REPORTS_NEED_TO_BE_INSERTED_IN_WORKPATH = True
    THERE_ARE_NEW_NEP_REPORTS_NEED_TO_BE_INSERTED_IN_ROHS = True
    # ver.1.1 add a if judge to void error if there is no NG report in workpath.
    # ver.1.2 add a check and a if to confrim that new rows of NG reports will be inserted in to the BI source data, as the NG reports will repeat in serverl days.
    print("================Updating the NEP inspection data.================")
    print("=       Ver.1.1                                                 =")
    print("=       2022/6/10                                               =")
    print("=       IoTG QC                                                 =")
    print("=  1.update NEP UPS_Acceptance_Defect_NEP_FY19- to file server  =")
    print("=  2.update RoHS保証運用管理表 to file server                   =")
    print("=  3.copy NEP reports to file server                            =")
    print("=  3.1.OMRON UPS INSPECTION REPORT                              =")    
    print("=  3.2.NGReport                                                 =")    
    print("=  3.3.SNログ                                                   =")    
    print("=================================================================")



# 1.1 download the BI data source
    if os.path.isfile(BI_data_source_online_path_name):
        os.system(copy_command_NEP_BI_source_data_from_online_to_local)
        print("Download the BI source data from file server to local workpath.Please wait.")
        timedown = 5
        while timedown :
            time.sleep(1)
            if os.path.isfile(BI_data_source_local_path_name):
                print("The BI source data has been downloaded from file server to local workpath successfully.")
                break
            elif timedown >=0:
                print("Please wait.")
                timedown = timedown - 1
            elif timedown == 0:
                print("The BI data can't be downloaded as network is abnormal.")
                exit()
    else:
        print("Error.Can't find the BI source data.")
        exit()

# 1.2 download the RoHS data source
    if os.path.isfile(NEP_RoHS_online_path_name):
        os.system(copy_command_NEP_RoHS_from_online_to_local)
        print("Download the RoHS source data from file server to local workpath.Please wait.")
        timedown = 5
        while timedown :
            time.sleep(1)
            if os.path.isfile(NEP_RoHS_local_path_name):
                print("The RoHS source data has been downloaded from file server to local workpath successfully.")
                break
            elif timedown >=0:
                print("Please wait.")
                timedown = timedown - 1
            elif timedown == 0:
                print("The RoHS data can't be downloaded as network is abnormal.")
                exit()
    else:
        print("Error.Can't find the RoHS source data.")
        exit()

# 2.1 open sheet "検査実績" of the BI data with pd
    df_BI_data_source_NEP_inspection_data_tmp = pd.DataFrame
    df_BI_data_source_NEP_inspection_data_tmp = pd.read_excel(BI_data_source_local_path_name,sheet_name='検査実績',header=0,engine='openpyxl')
    # delete BI source data Unnamed:23-25
    df_BI_data_source_NEP_inspection_data_tmp.drop(columns=['Unnamed: 23','Unnamed: 24','Unnamed: 25'], inplace=True)

# 2.2 open sheet "NG" of the BI data with pd
    df_BI_data_source_NEP_NG_data_tmp = pd.DataFrame
    df_BI_data_source_NEP_NG_data_tmp = pd.read_excel(BI_data_source_local_path_name,sheet_name='NG',header=0,engine='openpyxl')

# 2.2 open sheet 1 of the RoHS with pd
    df_RoHS_source_NEP_inspection_data_tmp = pd.DataFrame
    df_RoHS_source_NEP_inspection_data_tmp = pd.read_excel(NEP_RoHS_local_path_name,sheet_name='1',header=0,engine='openpyxl')

# print(df_RoHS_source_NEP_inspection_data_tmp)
#         NEP検査日    Omron PO   MODEL    QTY  ... BY50S BY80S  BY120S  BY75SW
# 0          NaT  J4P2482269   BY35S  216.0  ...   NaN   NaN     NaN     NaN
# 1          NaT  J4P2505838   BY35S  270.0  ...   NaN   NaN     NaN     NaN
# 2          NaT  J4P2505846   BY35S  270.0  ...   NaN   NaN     NaN     NaN

# 3.1 open all NEP inspection reports with pd and contact them(or it) to one file
    reports_local_path_name_NEP_inspection_data_list = []
    regEx = '*' + 'xlsx'
    reports_local_path_name_NEP_inspection_data_list = glob.glob(NEP_inspection_data_local_path_name + regEx, recursive=True)
    num_of_inspection_reports = len(glob.glob(NEP_inspection_data_local_path_name + regEx, recursive=True))
    print("%d NEP inspection report(s) data will be imported to the tmp file."%num_of_inspection_reports)

    df_local_NEP_inspection_data_report_tmp = pd.DataFrame
    df_local_NEP_inspection_data_reports_tmp = pd.DataFrame

    for num in range(0,num_of_inspection_reports):
        if num == 0:
            df_local_NEP_inspection_data_reports_tmp = pd.read_excel(reports_local_path_name_NEP_inspection_data_list[num], sheet_name="UPS INSPECTION REPORT",header=6,engine='openpyxl')
            print("--------report %d--------"%(num+1))
            print("%s "%reports_local_path_name_NEP_inspection_data_list[num])
        else:
            print("--------report %d--------"%(num+1))
            print("%s "%reports_local_path_name_NEP_inspection_data_list[num])
            df_local_NEP_inspection_data_report_tmp = pd.read_excel(reports_local_path_name_NEP_inspection_data_list[num],sheet_name="UPS INSPECTION REPORT",header=6,engine='openpyxl')
            df_local_NEP_inspection_data_reports_tmp = pd.concat([df_local_NEP_inspection_data_reports_tmp,df_local_NEP_inspection_data_report_tmp], axis=0)
    print("%d NEP report(s) data has be imported to the tmp file successfully."%num_of_inspection_reports)

    # check the rows of the NEP inspection reports is correct
    df_local_NEP_inspection_data_reports_tmp.dropna(axis=0,thresh=17,inplace=True)
    i = df_local_NEP_inspection_data_reports_tmp.shape[0]
    df_local_NEP_inspection_data_reports_tmp.dropna(axis=0, subset=['Date','OK'],inplace=True)
    j = df_local_NEP_inspection_data_reports_tmp.shape[0]
    if i == j:
        print("confirmation OK, NEP inspection reports data has been extracted successfully.")
    else:
        print("Error, NEP inspection reports.")
        exit()


# 3.2 open all NEP NG reports with pd and contact them(or it) to one file
    reports_local_path_name_NEP_NG_data_list = []
    regEx = '*' + 'xlsm'
    reports_local_path_name_NEP_NG_data_list = glob.glob(NEP_NG_data_local_path_name + regEx, recursive=True)
    num_of_NG_reports = len(glob.glob(NEP_NG_data_local_path_name + regEx, recursive=True))
    print("%d NEP NG report(s) data will be imported to the tmp file."%num_of_NG_reports)

    # ver.1.1 if NO_NEP_NG_REPORT == True, slip all the process to deal the NG reports.
    if reports_local_path_name_NEP_NG_data_list == [] :
        THERE_ARE_NEP_NG_REPORTS_IN_WORKPATH = False


    print(reports_local_path_name_NEP_NG_data_list)
    if THERE_ARE_NEP_NG_REPORTS_IN_WORKPATH == False:
        print("There is no NG report in workpath.")
    else:
        df_local_NEP_NG_data_report_tmp = pd.DataFrame
        df_local_NEP_NG_data_reports_tmp = pd.DataFrame

        for num in range(0,num_of_NG_reports):
            if num == 0:
                df_local_NEP_NG_data_reports_tmp = pd.read_excel(reports_local_path_name_NEP_NG_data_list[num], sheet_name=[0],header=3,engine='openpyxl')
                # if df_local_NEP_NG_data_reports_tmp is dict, change the type to DataFrame
                if isinstance(df_local_NEP_NG_data_reports_tmp, dict):
                    df_local_NEP_NG_data_reports_tmp = df_local_NEP_NG_data_reports_tmp[0]
                print("--------report %d--------"%(num+1))
                print("%s "%reports_local_path_name_NEP_NG_data_list[num])
            else:
                print("--------report %d--------"%(num+1))
                print("%s "%reports_local_path_name_NEP_NG_data_list[num])
                df_local_NEP_NG_data_report_tmp = pd.read_excel(reports_local_path_name_NEP_NG_data_list[num],sheet_name=[0],header=3,engine='openpyxl')
                # if df_local_NEP_NG_data_reports_tmp is dict, change the type to DataFrame
                if isinstance(df_local_NEP_NG_data_report_tmp, dict):
                    df_local_NEP_NG_data_report_tmp = df_local_NEP_NG_data_report_tmp[0]
                df_local_NEP_NG_data_reports_tmp = pd.concat([df_local_NEP_NG_data_reports_tmp,df_local_NEP_NG_data_report_tmp], axis=0)
        print("%d NEP NG report(s) data has be imported to the tmp file successfully."%num_of_NG_reports)
        print(df_local_NEP_NG_data_reports_tmp)


# check the rows of the NEP NG reports is correct
    if THERE_ARE_NEP_NG_REPORTS_IN_WORKPATH == True:
        df_local_NEP_NG_data_reports_tmp.dropna(axis=0,thresh=8,inplace=True)
        i = df_local_NEP_NG_data_reports_tmp.shape[0]
        df_local_NEP_NG_data_reports_tmp.dropna(axis=0, subset=['Model','Serial Number'],inplace=True)
        j = df_local_NEP_NG_data_reports_tmp.shape[0]
        if i == j:
            print("confirmation OK, NEP NG reports data has been extracted successfully.")
        else:
            print("Error, NEP NG reports.")
            exit()

        # ver.1.2 
        # column=9 is SN
        # df.drop(index=(df.loc[(df['table']=='sc')].index),inplace=True) NG-> index No. is 0, 0, 0
        # df.drop(df[(df.score < 50) & (df.score >20)].index, inplace=True) NG-> index No. is 0, 0, 0
        # df_new =  df.reset_index(drop=True) OK-> change index No.
        df_BI_data_source_NEP_NG_data_tmp_for_check = df_BI_data_source_NEP_NG_data_tmp.iloc[-50:, 9]
        df_new = df_local_NEP_NG_data_reports_tmp.reset_index(drop=True)
        df_local_NEP_NG_data_reports_tmp = df_new

        for i in df_local_NEP_NG_data_reports_tmp['Serial Number']:
           for j in  df_BI_data_source_NEP_NG_data_tmp_for_check:
                if i == j:
                    # Int64Index([0], dtype='int64')
                    # Int64Index([0, 0], dtype='int64')
                    # Int64Index([], dtype='int64')
                    df_local_NEP_NG_data_reports_tmp.drop(df_local_NEP_NG_data_reports_tmp[(df_local_NEP_NG_data_reports_tmp['Serial Number']==i)].index , inplace=True)

        if len(df_local_NEP_NG_data_reports_tmp) == 0:
            THERE_ARE_NEW_NEP_NG_REPORTS_NEED_TO_BE_INSERTED_IN_WORKPATH = False
            print("There is no new NG report will be inserted into the BI source file.")

    # THERE_ARE_NEP_NG_REPORTS_IN_WORKPATH = True
    # THERE_ARE_NEW_NEP_NG_REPORTS_NEED_TO_BE_INSERTED_IN_WORKPATH = True

        # print(NO_NEP_NG_REPORT)
        # print(df_local_NEP_NG_data_reports_tmp)



# 4.1　check 3. NEP inspection reports which is waiting to be inserted to the BI source data
    # get the value(Order No.) of the last 100 rows in BI data source "検査実績" sheet
    BI_source_data_Order_No_last_100_rows = df_BI_data_source_NEP_inspection_data_tmp.iloc[-100:,2]
    # get all value(Order No.) of the rows in inspection reports
    list_all_order_No_in_reports = df_local_NEP_inspection_data_reports_tmp.iloc[:,2]
    compare_two_columns_source_insert(BI_source_data_Order_No_last_100_rows, list_all_order_No_in_reports)

# 4.2 check 3. NEP NG reports which is waiting to be inserted to the BI source data
    if THERE_ARE_NEP_NG_REPORTS_IN_WORKPATH == True & THERE_ARE_NEW_NEP_NG_REPORTS_NEED_TO_BE_INSERTED_IN_WORKPATH == True:
        # get the value(Order No.) of the last 20 rows in BI data source "NG" sheet
        BI_source_data_Order_No_last_100_rows = df_BI_data_source_NEP_NG_data_tmp.iloc[-20:,9]
        # get all value(Order No.) of the rows in NG reports
        list_all_order_No_in_reports = df_local_NEP_NG_data_reports_tmp.iloc[:,9]
        compare_two_columns_source_insert(BI_source_data_Order_No_last_100_rows, list_all_order_No_in_reports)

# 4.3  check 3. which is waiting to be inserted into the RoHS source data
    # select the rows which represent model should be insert into the RoHS source data
    df_RoHS_source_NEP_inspection_data_tmp_selected = df_local_NEP_inspection_data_reports_tmp[(df_local_NEP_inspection_data_reports_tmp['Model']=='BY35S')|(df_local_NEP_inspection_data_reports_tmp['Model']=='BY50S')|(df_local_NEP_inspection_data_reports_tmp['Model']=='BY80S')|(df_local_NEP_inspection_data_reports_tmp['Model']=='BY120S')|(df_local_NEP_inspection_data_reports_tmp['Model']=='BY75SW')]
    print(df_RoHS_source_NEP_inspection_data_tmp_selected)
    print(len(df_RoHS_source_NEP_inspection_data_tmp_selected))
    if len(df_RoHS_source_NEP_inspection_data_tmp_selected) == 0:
        THERE_ARE_NEW_NEP_REPORTS_NEED_TO_BE_INSERTED_IN_ROHS = False

    if THERE_ARE_NEW_NEP_REPORTS_NEED_TO_BE_INSERTED_IN_ROHS == True:
        # get the value(Order No.) of the last 100 rows in RoHS source data 
        RoHS_Order_No_last_100_rows = df_RoHS_source_NEP_inspection_data_tmp.iloc[-100:,1]
        # get all value(Order No.) of the rows in reports
        list_all_order_No_in_reports_selected = df_RoHS_source_NEP_inspection_data_tmp_selected.iloc[:,2]
        compare_two_columns_source_insert(RoHS_Order_No_last_100_rows, list_all_order_No_in_reports_selected)


# 5.insert 3. to BI source data
# 5.1 use the column of Date in NEP inspction reports to create yearmonth series
    NEP_inspection_data_Date_list = df_local_NEP_inspection_data_reports_tmp.iloc[:,1]
    NEP_inspection_data_yearmonth_list = []
    for l in NEP_inspection_data_Date_list:
        # use year - 2000 to get the double-digit to express year
        year_str =str(l.year - 2000)
        if l.month >= 10:
            month_str = str(l.month)
        else:
            month_str =  "0" + str(l.month)
        yearmonth_int = int(year_str + month_str)
        NEP_inspection_data_yearmonth_list.append(yearmonth_int)
    # add the yearmonth series to BI source data as the last column from the right
    df_local_NEP_inspection_data_reports_tmp['yearmonth'] = NEP_inspection_data_yearmonth_list

# 5.2 use the column of Date in NEP NG reports to create 年月 series and 年月2 series
    if THERE_ARE_NEP_NG_REPORTS_IN_WORKPATH == True & THERE_ARE_NEW_NEP_NG_REPORTS_NEED_TO_BE_INSERTED_IN_WORKPATH == True:
        NEP_NG_data_Date_list = df_local_NEP_NG_data_reports_tmp.iloc[:,0]
        NEP_NG_data_nengetu_list = []
        NEP_NG_data_nengetu2_list = []
        for l in NEP_NG_data_Date_list:
            # use year - 2000 to get the double-digit to express year
            year_str =str(l.year - 2000)
            if l.month >= 10:
                month_str = str(l.month)
            else:
                month_str =  "0" + str(l.month)
            yearmonth_int = int(year_str + month_str)
            NEP_NG_data_nengetu_list.append(yearmonth_int)
        # add the yearmonth series to BI source data as the last column from the right
        df_local_NEP_NG_data_reports_tmp['年月'] = NEP_NG_data_nengetu_list



# 6 insert reports
    # 6.1 insert inspection reports
    # confirm the number of the rows in BI data source
    last_row_num = df_BI_data_source_NEP_inspection_data_tmp.shape[0]
    write_excel_xlsx_append_xlwings(BI_data_source_local_path_name, "検査実績",last_row_num, df_local_NEP_inspection_data_reports_tmp,1)

    # 6.2 insert inspection reports
    # confirm the number of the rows in BI data source
    if THERE_ARE_NEP_NG_REPORTS_IN_WORKPATH == True & THERE_ARE_NEW_NEP_NG_REPORTS_NEED_TO_BE_INSERTED_IN_WORKPATH == True:
        last_row_num = df_BI_data_source_NEP_NG_data_tmp.shape[0]
        write_excel_xlsx_append_xlwings(BI_data_source_local_path_name, "NG",last_row_num, df_local_NEP_NG_data_reports_tmp,1)
   
    # 6.3 insert RoHS with inspection reports
    if THERE_ARE_NEW_NEP_REPORTS_NEED_TO_BE_INSERTED_IN_ROHS == True:
        df_RoHS_source_NEP_inspection_data_tmp_selected.drop(df_RoHS_source_NEP_inspection_data_tmp_selected.columns[0], axis=1, inplace=True)
        df_RoHS_source_NEP_inspection_data_tmp_selected.drop(df_RoHS_source_NEP_inspection_data_tmp_selected.columns[4:], axis=1, inplace=True)
        # list_all_order_No_in_reports_selected.drop([columns=[1], axis=1, inplace=True])
        last_row_num = df_RoHS_source_NEP_inspection_data_tmp.shape[0]
        write_excel_xlsx_append_xlwings(NEP_RoHS_local_path_name, "1", last_row_num, df_RoHS_source_NEP_inspection_data_tmp_selected,1)
    else:
        print("There is no update to the RoHS source file.")

# 7 copy the source data and reports to online.
    if Copy_To_Online:
        # BI source data
        if os.path.isfile(BI_data_source_local_path_name):
            os.system(copy_command_NEP_BI_source_data_from_local_to_online)
            print("Upload the BI source data from  local workpath to file server.Please wait.")
            timedown = 5
            while timedown :
                time.sleep(1)
                if os.path.isfile(BI_data_source_online_path_name):
                    print("The BI source data has been uploaded from local workpath to file server successfully.")
                    os.remove(BI_data_source_local_path_name)
                    break
                elif timedown >=0:
                    print("Please wait.")
                    timedown = timedown - 1
                elif timedown == 0:
                    print("The BI data can't be uploaded as network is abnormal.")
                    exit()
        else:
            print("Error.The BI source data has not been aploaded successsfully.")
            exit()
        # RoHS source data
        if os.path.isfile(NEP_RoHS_local_path_name):
            os.system(copy_command_NEP_RoHS_from_local_to_online)
            print("Upload the RoHS source data from  local workpath to file server.Please wait.")
            timedown = 5
            while timedown :
                time.sleep(1)
                if os.path.isfile(NEP_RoHS_online_path_name):
                    print("The RoHS source data has been uploaded from local workpath to file server successfully.")
                    os.remove(NEP_RoHS_local_path_name)
                    break
                elif timedown >=0:
                    print("Please wait.")
                    timedown = timedown - 1
                elif timedown == 0:
                    print("The RoHS data can't be uploaded as network is abnormal.")
                    exit()
        else:
            print("Error.The RoHS source data has not been aploaded successsfully.")
            exit()
        # NEP inspection reports

        # reports_local_path_name_NEP_inspection_data_list
        print("%d NEP inspection report(s) data will be copyed to file server."%len(reports_local_path_name_NEP_inspection_data_list))
        for i in reports_local_path_name_NEP_inspection_data_list:
            copy_command_NEP_inspection_data_from_local_to_online =  'copy' + ' ' + '"' + i  + '"'  +  ' ' + '"' + NEP_inspection_data_online_path +'"'
            os.system(copy_command_NEP_inspection_data_from_local_to_online)
            os.remove(i)
            print("%s has been uploaded to file server."%i)
        print("All NEP inspection data have been uploaded from local workpath to file server.")
        
        if THERE_ARE_NEP_NG_REPORTS_IN_WORKPATH == True:
            # reports_local_path_name_NEP_NG_data_list
            print("%d NEP NG report(s) data will be copyed to file server."%len(reports_local_path_name_NEP_NG_data_list))
            for i in reports_local_path_name_NEP_NG_data_list:
                copy_command_NEP_NG_data_from_local_to_online =  'copy' + ' ' + '"' + i  + '"'  +  ' ' + '"' + NEP_NG_data_online_path +'"'
                os.system(copy_command_NEP_NG_data_from_local_to_online)
                os.remove(i)
                print("%s has been uploaded to file server."%i)
            print("All NEP NG data (.xlsm) have been uploaded from local workpath to file server.")

            # NG reports .pdf
            reports_local_path_name_NEP_NG_data_list = []
            regEx = '*' + 'pdf'
            reports_local_path_name_NEP_NG_data_list = glob.glob(NEP_NG_data_local_path_name + regEx, recursive=True)
            print("%d NEP NG report(s) data will be copyed to file server."%len(glob.glob(NEP_NG_data_local_path_name + regEx, recursive=True)))
            for i in reports_local_path_name_NEP_NG_data_list:
                copy_command_NEP_NG_data_from_local_to_online =  'copy' + ' ' + '"' + i  + '"'  +  ' ' + '"' + NEP_NG_data_online_path +'"'
                os.system(copy_command_NEP_NG_data_from_local_to_online)
                os.remove(i)
                print("%s has been uploaded to file server."%i)
            print("All NEP NG data (.pdf) have been uploaded from local workpath to file server.")

        # SNログ
        reports_local_path_name_NEP_SN_data_list = []
        regEx = '*' + 'xlsx'
        reports_local_path_name_NEP_SN_data_list = glob.glob(NEP_SN_data_local_path_name + regEx, recursive=True)
        print("%d NEP SN report(s) data will be copyed to file server."%len(glob.glob(NEP_SN_data_local_path_name + regEx, recursive=True)))
        for i in reports_local_path_name_NEP_SN_data_list:
            copy_command_NEP_SN_data_from_local_to_online =  'copy' + ' ' + '"' + i  + '"'  +  ' ' + '"' + NEP_SN_data_online_path +'"'
            os.system(copy_command_NEP_SN_data_from_local_to_online)
            os.remove(i)
            print("%s has been uploaded to file server."%i)
        print("All NEP SN data have been uploaded from local workpath to file server.")


if __name__ == '__main__':
    main()
